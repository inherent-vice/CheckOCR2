"""Extract date/rate OCR fixture crops from copied CouponCheck full-area images."""

from __future__ import annotations

import argparse
import csv
import json
import sys
import tempfile
from collections.abc import Sequence
from pathlib import Path
from typing import Any

from PIL import Image

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from checkocr2.ocr_text import clean_date_text, clean_rate_text  # noqa: E402
from checkocr2.paths import sanitize_filename  # noqa: E402

DEFAULT_OUTPUT_DIR = Path(".analysis_tmp/ocr_crops")
DEFAULT_CSV_NAME = "ground_truth.csv"
DEFAULT_DATE_ROI = "0.62,0,0.80,0.075"
DEFAULT_RATE_ROI = "0.80,0,1.0,0.075"
DEFAULT_FULL_DATE_ROI = "0.63,0.158,0.79,0.23"
DEFAULT_FULL_RATE_ROI = "0.81,0.158,0.985,0.23"
FIELDNAMES = ["crop_path", "field", "expected_text", "source_run", "notes"]


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--day-dir",
        type=Path,
        nargs="+",
        required=True,
        help="One or more copied real-data day directories, e.g. .analysis_tmp/real_data/20260513.",
    )
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--csv-name", default=DEFAULT_CSV_NAME)
    parser.add_argument("--date-roi", default=DEFAULT_DATE_ROI)
    parser.add_argument("--rate-roi", default=DEFAULT_RATE_ROI)
    parser.add_argument("--full-date-roi", default=DEFAULT_FULL_DATE_ROI)
    parser.add_argument("--full-rate-roi", default=DEFAULT_FULL_RATE_ROI)
    parser.add_argument("--layout", choices=("auto", "cropped", "full"), default="auto")
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--allow-unsafe-output",
        action="store_true",
        help="Allow output outside .analysis_tmp, tests/fixtures/ocr_crops, or temp.",
    )
    return parser.parse_args(argv)


def extract_real_data_ocr_fixtures(
    *,
    day_dir: Path,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    csv_name: str = DEFAULT_CSV_NAME,
    date_roi: str = DEFAULT_DATE_ROI,
    rate_roi: str = DEFAULT_RATE_ROI,
    full_date_roi: str = DEFAULT_FULL_DATE_ROI,
    full_rate_roi: str = DEFAULT_FULL_RATE_ROI,
    layout: str = "auto",
    limit: int = 0,
    overwrite: bool = False,
    dry_run: bool = False,
    allow_unsafe_output: bool = False,
) -> dict[str, Any]:
    return extract_real_data_ocr_fixture_days(
        day_dirs=[day_dir],
        output_dir=output_dir,
        csv_name=csv_name,
        date_roi=date_roi,
        rate_roi=rate_roi,
        full_date_roi=full_date_roi,
        full_rate_roi=full_rate_roi,
        layout=layout,
        limit=limit,
        overwrite=overwrite,
        dry_run=dry_run,
        allow_unsafe_output=allow_unsafe_output,
    )


def extract_real_data_ocr_fixture_days(
    *,
    day_dirs: Sequence[Path],
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    csv_name: str = DEFAULT_CSV_NAME,
    date_roi: str = DEFAULT_DATE_ROI,
    rate_roi: str = DEFAULT_RATE_ROI,
    full_date_roi: str = DEFAULT_FULL_DATE_ROI,
    full_rate_roi: str = DEFAULT_FULL_RATE_ROI,
    layout: str = "auto",
    limit: int = 0,
    overwrite: bool = False,
    dry_run: bool = False,
    allow_unsafe_output: bool = False,
) -> dict[str, Any]:
    if not day_dirs:
        raise ValueError("at least one day directory is required")

    resolved_day_dirs = [day_dir.resolve() for day_dir in day_dirs]
    output_dir = output_dir.resolve()
    validate_layout(layout)
    roi_by_layout = {
        "cropped": {
            "date": parse_roi(date_roi, "date"),
            "rate": parse_roi(rate_roi, "rate"),
        },
        "full": {
            "date": parse_roi(full_date_roi, "full date"),
            "rate": parse_roi(full_rate_roi, "full rate"),
        },
    }
    for day_dir in resolved_day_dirs:
        validate_inputs(
            day_dir=day_dir,
            output_dir=output_dir,
            csv_name=csv_name,
            allow_unsafe_output=allow_unsafe_output,
        )

    csv_path = output_dir / csv_name
    if csv_path.exists() and not overwrite and not dry_run:
        raise FileExistsError(f"fixture CSV already exists: {csv_path}")

    planned_crops: list[dict[str, Any]] = []
    rows: list[dict[str, str]] = []
    existing_outputs: list[str] = []
    missing_images: list[str] = []
    skipped_blank_expected = 0
    row_count_from_workbooks = 0
    day_summaries: list[dict[str, Any]] = []

    for day_dir in resolved_day_dirs:
        day_summary = plan_day_fixture_rows(
            day_dir=day_dir,
            output_dir=output_dir,
            roi_by_layout=roi_by_layout,
            layout=layout,
            rows=rows,
            planned_crops=planned_crops,
            existing_outputs=existing_outputs,
            missing_images=missing_images,
            overwrite=overwrite,
            dry_run=dry_run,
            limit=limit,
        )
        skipped_blank_expected += day_summary["skipped_blank_expected_count"]
        row_count_from_workbooks += day_summary["row_count_from_workbook"]
        day_summaries.append(day_summary)

    if existing_outputs:
        raise FileExistsError("fixture crop already exists: " + ", ".join(existing_outputs[:3]))
    if not rows:
        raise ValueError("no fixture rows could be generated from updated workbook and images")

    summary = {
        "status": "planned" if dry_run else "ready",
        "day_dirs": [str(day_dir) for day_dir in resolved_day_dirs],
        "days": day_summaries,
        "output_dir": str(output_dir),
        "fixture_csv": str(csv_path),
        "total_cases": len(rows),
        "field_counts": count_fields(rows),
        "row_count_from_workbook": row_count_from_workbooks,
        "missing_image_count": len(missing_images),
        "missing_images": missing_images[:20],
        "skipped_blank_expected_count": skipped_blank_expected,
        "layout": layout,
        "roi": {
            layout_name: {
                field: format_roi(roi)
                for field, roi in field_rois.items()
            }
            for layout_name, field_rois in roi_by_layout.items()
        },
        "dry_run": dry_run,
    }
    if dry_run:
        return summary

    output_dir.mkdir(parents=True, exist_ok=True)
    if overwrite:
        remove_existing_fixture_outputs(output_dir, rows)
    for planned in planned_crops:
        crop_and_save(
            source=planned["source"],
            destination=planned["destination"],
            roi=planned["roi"],
        )
    write_fixture_csv(csv_path, rows)
    write_manifest(output_dir / "fixture_manifest.json", summary)
    return summary


def plan_day_fixture_rows(
    *,
    day_dir: Path,
    output_dir: Path,
    roi_by_layout: dict[str, dict[str, tuple[float, float, float, float]]],
    layout: str,
    rows: list[dict[str, str]],
    planned_crops: list[dict[str, Any]],
    existing_outputs: list[str],
    missing_images: list[str],
    overwrite: bool,
    dry_run: bool,
    limit: int,
) -> dict[str, Any]:
    updated_workbook = find_updated_workbook(day_dir)
    image_dir = day_dir / "images"
    expected_rows = load_expected_rows(updated_workbook)
    if limit > 0:
        expected_rows = expected_rows[:limit]

    start_row_count = len(rows)
    start_missing_count = len(missing_images)
    skipped_blank_expected = 0

    for expected in expected_rows:
        code = expected["code"]
        image_path = image_dir / f"{code}.png"
        if not image_path.exists():
            missing_images.append(f"{day_dir.name}/{code}")
            continue
        image_layout = layout_for_image(image_path, layout)
        for field in ("date", "rate"):
            expected_text = normalize_expected(field, expected.get(field, ""))
            if not expected_text:
                skipped_blank_expected += 1
                continue
            roi = roi_by_layout[image_layout][field]
            output_name = fixture_filename(len(rows) + 1, day_dir.name, code, field)
            destination = output_dir / output_name
            if destination.exists() and not overwrite and not dry_run:
                existing_outputs.append(str(destination))
            planned_crops.append(
                {
                    "source": image_path,
                        "destination": destination,
                        "field": field,
                        "roi": roi,
                }
            )
            rows.append(
                {
                    "crop_path": output_name,
                    "field": field,
                    "expected_text": expected_text,
                    "source_run": day_dir.name,
                        "notes": (
                            f"source={day_dir.name}/images/{code}.png; "
                            "expected_from_updated_workbook; "
                            f"layout={image_layout}; "
                            f"roi={format_roi(roi)}"
                        ),
                }
            )

    return {
        "day_dir": str(day_dir),
        "updated_workbook": str(updated_workbook),
        "image_dir": str(image_dir),
        "row_count_from_workbook": len(expected_rows),
        "total_cases": len(rows) - start_row_count,
        "field_counts": count_fields(rows[start_row_count:]),
        "missing_image_count": len(missing_images) - start_missing_count,
        "skipped_blank_expected_count": skipped_blank_expected,
    }


def parse_roi(value: str, field: str) -> tuple[float, float, float, float]:
    parts = [part.strip() for part in value.split(",")]
    if len(parts) != 4:
        raise ValueError(f"{field} ROI must contain four comma-separated fractions")
    try:
        left, top, right, bottom = (float(part) for part in parts)
    except ValueError as exc:
        raise ValueError(f"{field} ROI must contain numeric fractions") from exc
    if not (0 <= left < right <= 1 and 0 <= top < bottom <= 1):
        raise ValueError(f"{field} ROI fractions must satisfy 0 <= left < right <= 1")
    return left, top, right, bottom


def validate_layout(layout: str) -> None:
    if layout not in {"auto", "cropped", "full"}:
        raise ValueError("layout must be auto, cropped, or full")


def layout_for_image(image_path: Path, layout: str) -> str:
    validate_layout(layout)
    if layout != "auto":
        return layout
    with Image.open(image_path) as image:
        width, height = image.size
    return "cropped" if width >= 800 and height >= 470 else "full"


def validate_inputs(
    *,
    day_dir: Path,
    output_dir: Path,
    csv_name: str,
    allow_unsafe_output: bool,
) -> None:
    if not day_dir.exists() or not day_dir.is_dir():
        raise FileNotFoundError(f"day directory not found: {day_dir}")
    image_dir = day_dir / "images"
    if not image_dir.exists() or not image_dir.is_dir():
        raise FileNotFoundError(f"image directory not found: {image_dir}")

    csv_path = Path(csv_name)
    if csv_path.is_absolute() or csv_path.name != csv_name or csv_path.parent != Path("."):
        raise ValueError(f"csv_name must be a filename, not a path: {csv_name}")
    if csv_path.suffix.lower() != ".csv":
        raise ValueError(f"csv_name must end with .csv: {csv_name}")
    if output_dir == image_dir or is_relative_to(output_dir, image_dir):
        raise ValueError("output_dir must not be the source image directory or one of its children")

    allowed_roots = [
        (ROOT / ".analysis_tmp").resolve(),
        (ROOT / "tests" / "fixtures" / "ocr_crops").resolve(),
        Path(tempfile.gettempdir()).resolve(),
    ]
    if any(output_dir == allowed or is_relative_to(output_dir, allowed) for allowed in allowed_roots):
        return
    if not allow_unsafe_output:
        raise ValueError(
            "output_dir must be under .analysis_tmp, tests/fixtures/ocr_crops, or temp; "
            "pass --allow-unsafe-output only for deliberate local experiments"
        )


def find_updated_workbook(day_dir: Path) -> Path:
    candidates = sorted(day_dir.glob("*_updated.xlsx"))
    if not candidates:
        raise FileNotFoundError(f"updated workbook not found in: {day_dir}")
    if len(candidates) > 1:
        raise ValueError("multiple updated workbooks found: " + ", ".join(path.name for path in candidates))
    return candidates[0]


def load_expected_rows(updated_workbook: Path) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(updated_workbook, read_only=True, data_only=True)
    worksheet = workbook.active
    header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        raise ValueError(f"updated workbook has no header: {updated_workbook}")
    columns = header_columns(header)
    rows: list[dict[str, str]] = []
    for raw_row in worksheet.iter_rows(min_row=2, values_only=True):
        code = string_cell(raw_row[columns["code"]])
        if not code:
            continue
        rows.append(
            {
                "code": code,
                "date": string_cell(raw_row[columns["date"]]),
                "rate": string_cell(raw_row[columns["rate"]]),
            }
        )
    return rows


def header_columns(header: tuple[Any, ...]) -> dict[str, int]:
    normalized = {string_cell(value): index for index, value in enumerate(header)}
    aliases = {
        "code": ("종목코드",),
        "date": ("날짜",),
        "rate": ("금리", "표면금리"),
    }
    columns: dict[str, int] = {}
    for key, names in aliases.items():
        for name in names:
            if name in normalized:
                columns[key] = normalized[name]
                break
        if key not in columns:
            raise ValueError(f"updated workbook missing required column: {names[0]}")
    return columns


def string_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_expected(field: str, text: str) -> str:
    if field == "date":
        return clean_date_text(text)
    if field == "rate":
        return clean_rate_text(text)
    return text.strip()


def fixture_filename(index: int, day: str, code: str, field: str) -> str:
    return f"{index:04d}_{sanitize_filename(day)}_{sanitize_filename(code)}_{field}.png"


def crop_and_save(
    *,
    source: Path,
    destination: Path,
    roi: tuple[float, float, float, float],
) -> None:
    with Image.open(source) as image:
        width, height = image.size
        box = roi_to_box(roi, width, height)
        cropped = image.convert("RGB").crop(box)
        if cropped.width <= 0 or cropped.height <= 0:
            raise ValueError(f"empty crop for {source}: {box}")
        cropped.save(destination)


def roi_to_box(
    roi: tuple[float, float, float, float],
    width: int,
    height: int,
) -> tuple[int, int, int, int]:
    left, top, right, bottom = roi
    return (
        max(0, min(width - 1, int(round(width * left)))),
        max(0, min(height - 1, int(round(height * top)))),
        max(1, min(width, int(round(width * right)))),
        max(1, min(height, int(round(height * bottom)))),
    )


def write_fixture_csv(csv_path: Path, rows: list[dict[str, str]]) -> None:
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)


def write_manifest(path: Path, summary: dict[str, Any]) -> None:
    path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def remove_existing_fixture_outputs(output_dir: Path, rows: list[dict[str, str]]) -> None:
    for row in rows:
        crop_path = output_dir / row["crop_path"]
        if crop_path.exists():
            crop_path.unlink()
    manifest = output_dir / "fixture_manifest.json"
    if manifest.exists():
        manifest.unlink()


def count_fields(rows: list[dict[str, str]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        field = row["field"]
        counts[field] = counts.get(field, 0) + 1
    return dict(sorted(counts.items()))


def format_roi(roi: tuple[float, float, float, float]) -> str:
    return ",".join(f"{value:g}" for value in roi)


def is_relative_to(path: Path, parent: Path) -> bool:
    try:
        path.relative_to(parent)
    except ValueError:
        return False
    return True


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        summary = extract_real_data_ocr_fixture_days(
            day_dirs=args.day_dir,
            output_dir=args.output_dir,
            csv_name=args.csv_name,
            date_roi=args.date_roi,
            rate_roi=args.rate_roi,
            full_date_roi=args.full_date_roi,
            full_rate_roi=args.full_rate_roi,
            layout=args.layout,
            limit=args.limit,
            overwrite=args.overwrite,
            dry_run=args.dry_run,
            allow_unsafe_output=args.allow_unsafe_output,
        )
    except (
        FileExistsError,
        FileNotFoundError,
        ValueError,
        OSError,
    ) as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 2
    print(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
