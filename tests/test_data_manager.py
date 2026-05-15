from __future__ import annotations

import logging
import queue
from zipfile import ZipFile

import pandas as pd
import pytest
from openpyxl import load_workbook

from checkocr2 import data_manager as data_manager_module

CODE = "\uc885\ubaa9\ucf54\ub4dc"
NAME = "\uc885\ubaa9\uba85"
DATE = "\ub0a0\uc9dc"
RATE = "\uae08\ub9ac"
STATUS = "\uc0c1\ud0dc"
WAITING = "\ub300\uae30 \uc911"
DONE = "\uc644\ub8cc"
STOPPED = "\uc911\ub2e8\ub428"


def make_data_manager(ocr_module):
    events = queue.Queue()
    manager = ocr_module.DataManager(
        app_ref=None,
        logger=logging.getLogger("tests.data_manager"),
        message_queue=events,
    )
    return manager, events


class FakeVar:
    def __init__(self, value):
        self.value = value

    def get(self):
        return self.value


def test_load_excel_to_grid_accepts_english_code_and_name_columns(ocr_module, tmp_path):
    workbook_path = tmp_path / "input.xlsx"
    pd.DataFrame(
        [
            {"code": "A001", "name": "Alpha"},
            {"code": "B002", "name": "Beta"},
        ],
    ).to_excel(workbook_path, index=False)

    manager, _events = make_data_manager(ocr_module)

    assert ocr_module.DataManager.__module__ == "checkocr2.data_manager"
    assert manager.load_excel_to_grid_data(workbook_path) == 2
    assert manager.excel_data == [
        {CODE: "A001", NAME: "Alpha", DATE: "", RATE: "", STATUS: WAITING},
        {CODE: "B002", NAME: "Beta", DATE: "", RATE: "", STATUS: WAITING},
    ]


def test_load_excel_to_grid_handles_corrupt_xlsx_without_raising(ocr_module, tmp_path):
    workbook_path = tmp_path / "corrupt.xlsx"
    workbook_path.write_bytes(b"PK-not-a-valid-xlsx")
    manager, events = make_data_manager(ocr_module)

    assert manager.load_excel_to_grid_data(workbook_path) == 0
    assert manager.excel_data == []
    queued = list(events.queue)
    assert queued
    assert queued[-1][0] == "error_messagebox"


def test_load_excel_to_grid_handles_zip_that_is_not_workbook(ocr_module, tmp_path):
    workbook_path = tmp_path / "not_workbook.xlsx"
    with ZipFile(workbook_path, "w") as workbook_zip:
        workbook_zip.writestr("not_workbook.txt", "hello")
    manager, events = make_data_manager(ocr_module)

    assert manager.load_excel_to_grid_data(workbook_path) == 0
    assert manager.excel_data == []
    queued = list(events.queue)
    assert queued
    assert queued[-1][0] == "error_messagebox"


def test_export_grid_to_excel_writes_updated_workbook(ocr_module, tmp_path):
    manager, events = make_data_manager(ocr_module)
    manager.excel_data = [
        {CODE: "A001", NAME: "Alpha", DATE: "2024/05/01", RATE: "3.500", STATUS: DONE},
        {CODE: "B002", NAME: "Beta", DATE: "", RATE: "", STATUS: STOPPED},
    ]

    output_dir = tmp_path / "output"
    output_dir.mkdir()
    exported_path = manager.export_grid_to_excel_data(output_dir, str(tmp_path / "source.xlsx"))

    assert exported_path.exists()
    assert exported_path.name == "source_updated.xlsx"
    worksheet = load_workbook(exported_path, data_only=True)["OCR_Results"]

    assert [cell.value for cell in worksheet[1]] == [CODE, NAME, DATE, RATE, STATUS]
    assert worksheet.cell(row=2, column=1).value == "A001"
    assert worksheet.cell(row=2, column=2).value == "Alpha"
    assert worksheet.cell(row=2, column=3).value.strftime("%Y/%m/%d") == "2024/05/01"
    assert worksheet.cell(row=2, column=3).number_format == "yyyy/mm/dd"
    assert worksheet.cell(row=2, column=4).value == 3.5
    assert worksheet.cell(row=2, column=4).number_format == "0.0000"
    assert worksheet.cell(row=2, column=5).value == DONE
    assert worksheet.cell(row=3, column=3).value in ("", None)
    assert worksheet.cell(row=3, column=4).value in ("", None)
    assert worksheet.cell(row=3, column=5).value in ("", None)
    assert any(event[0] == "log" and event[2] == "SUCCESS" for event in list(events.queue))


def test_export_grid_to_excel_uses_app_rate_decimal_places(ocr_module, monkeypatch, tmp_path):
    manager, _events = make_data_manager(ocr_module)
    manager.app = type("App", (), {"rate_decimal_places": FakeVar(2)})()
    manager.excel_data = [
        {CODE: "A001", NAME: "Alpha", DATE: "2024/05/01", RATE: "3.5", STATUS: DONE},
    ]
    captured = {}

    def capture_export_grid_rows(rows, output_path, *, rate_decimal_places=4):
        captured["rate_decimal_places"] = rate_decimal_places
        output_path.write_text("ok", encoding="utf-8")
        return output_path

    monkeypatch.setattr(data_manager_module, "export_grid_rows", capture_export_grid_rows)

    manager.export_grid_to_excel_data(tmp_path, str(tmp_path / "source.xlsx"))

    assert captured["rate_decimal_places"] == 2


def test_data_manager_row_edit_delete_clear_and_empty_export_contracts(ocr_module, tmp_path):
    manager, events = make_data_manager(ocr_module)
    manager.excel_data = [
        {CODE: "A001", NAME: "Alpha", DATE: "", RATE: "", STATUS: WAITING},
        {CODE: "B002", NAME: "Beta", DATE: "", RATE: "", STATUS: WAITING},
    ]
    manager.current_processing_index = 1

    assert manager.update_grid_cell_data(0, RATE, "3.500") is True
    assert manager.update_grid_cell_data(99, RATE, "4.000") is False
    manager.delete_rows_data([0])
    assert manager.excel_data[0][CODE] == "B002"
    manager.clear_all_data_internal()
    assert manager.excel_data == []
    assert manager.current_processing_index == -1
    assert manager.export_grid_to_excel_data(tmp_path, str(tmp_path / "source.xlsx")) is None
    assert list(events.queue)[-1] == ("log", "내보낼 데이터가 없습니다.", "INFO")


def test_paste_from_clipboard_data_queues_errors_without_mutating_rows(ocr_module):
    manager, events = make_data_manager(ocr_module)
    manager.excel_data = [{CODE: "A001", NAME: "Alpha", DATE: "", RATE: "", STATUS: WAITING}]

    assert manager.paste_from_clipboard_data(None) == 0
    assert manager.excel_data == [{CODE: "A001", NAME: "Alpha", DATE: "", RATE: "", STATUS: WAITING}]
    queued = list(events.queue)
    assert queued[-1][0] == "error_messagebox"
    assert queued[-1][1] == "붙여넣기 중 오류"


def test_export_grid_to_excel_propagates_writer_failures(ocr_module, monkeypatch, tmp_path):
    manager, _events = make_data_manager(ocr_module)
    manager.excel_data = [
        {CODE: "A001", NAME: "Alpha", DATE: "2024/05/01", RATE: "3.500", STATUS: DONE},
    ]

    def fail_export_grid_rows(rows, output_path, *, rate_decimal_places=4):
        raise PermissionError("locked workbook")

    monkeypatch.setattr(data_manager_module, "export_grid_rows", fail_export_grid_rows)

    with pytest.raises(PermissionError, match="locked workbook"):
        manager.export_grid_to_excel_data(tmp_path, str(tmp_path / "source.xlsx"))
