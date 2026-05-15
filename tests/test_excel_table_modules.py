from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import load_workbook

from checkocr2.events import GridUpdate, parse_legacy_grid_update
from checkocr2.excel_io import export_grid_rows, load_grid_rows, resolve_columns
from checkocr2.exceptions import ExcelIOError
from checkocr2.models import (
    CODE_COL,
    DATE_COL,
    NAME_COL,
    RATE_COL,
    STATUS_COL,
    STATUS_DONE,
    STATUS_ERROR_PROCESSING,
    STATUS_ERROR_SKIPPED,
    STATUS_PROCESSING,
    STATUS_STOPPED,
    STATUS_WAITING,
)
from checkocr2.table_model import (
    ClipboardSelection,
    GridStatusSummary,
    GridUpdateResult,
    apply_grid_update,
    delete_rows,
    format_grid_progress_text,
    format_grid_status_text,
    grid_cell_text,
    grid_row_tags,
    grid_row_values,
    rates_for_copy,
    row_for_copy,
    rows_for_copy,
    rows_for_export,
    rows_from_clipboard,
    status_is_error,
    summarize_grid_rows,
)


def test_resolve_columns_accepts_korean_and_english_aliases():
    resolved, missing = resolve_columns(["code", "회사명"])

    assert missing == []
    assert resolved[CODE_COL] == "code"
    assert resolved[NAME_COL] == "회사명"


def test_table_model_clipboard_delete_copy_and_export_status():
    rows = rows_from_clipboard("A001\tAlpha\nB002\tBeta")

    assert row_for_copy(rows[0]) == "A001\tAlpha\t\t\t대기 중"
    assert rows_for_copy(rows, [1, 99, 0]) == ClipboardSelection(
        text="B002\tBeta\t\t\t대기 중\nA001\tAlpha\t\t\t대기 중",
        count=2,
    )

    rows[0][RATE_COL] = "3.500"
    assert rates_for_copy(rows, [0, 99, 1]) == ClipboardSelection(text="3.500\n", count=2)

    rows[1][STATUS_COL] = STATUS_STOPPED
    export_rows = rows_for_export(rows)
    assert export_rows[1][STATUS_COL] == ""

    delete_rows(rows, [0])
    assert rows[0][CODE_COL] == "B002"


def test_grid_status_summary_and_labels_preserve_gui_text():
    rows = [
        {STATUS_COL: STATUS_DONE},
        {STATUS_COL: STATUS_WAITING},
        {STATUS_COL: STATUS_ERROR_PROCESSING},
        {STATUS_COL: "금리 없음"},
        {STATUS_COL: STATUS_ERROR_SKIPPED},
    ]

    summary = summarize_grid_rows(rows)

    assert summary == GridStatusSummary(total=5, completed=1, waiting=1, errors=3)
    assert summary.progress_percent == 20.0
    assert format_grid_status_text(summary) == "총 5행 | 완료: 1 | 대기: 1 | 오류: 3"
    assert format_grid_progress_text(summary) == "진행률: 20.0%"
    assert status_is_error("정상") is False


def test_grid_row_values_and_tags_preserve_treeview_contract():
    row = {
        CODE_COL: "A001",
        NAME_COL: "Alpha",
        DATE_COL: "2026/05/11",
        RATE_COL: "3.500",
        STATUS_COL: STATUS_WAITING,
    }

    assert grid_row_values(row) == ("A001", "Alpha", "2026/05/11", "3.500", STATUS_WAITING)
    assert grid_row_tags(row, row_index=0, current_processing_index=0, is_running=True) == ("processing",)
    assert grid_row_tags({STATUS_COL: STATUS_DONE}, row_index=0, current_processing_index=0, is_running=True) == (
        "completed",
    )
    assert grid_row_tags({STATUS_COL: "금리 없음"}, row_index=0, current_processing_index=0, is_running=True) == (
        "error",
    )
    assert grid_row_tags(row, row_index=1, current_processing_index=0, is_running=True) == ()

    with pytest.raises(KeyError):
        grid_row_values({STATUS_COL: STATUS_WAITING})


def test_apply_grid_update_preserves_legacy_tuple_behavior():
    rows = rows_from_clipboard("A001\tAlpha\nB002\tBeta")

    processing = apply_grid_update(rows, parse_legacy_grid_update(("processing", 1)))

    assert processing == GridUpdateResult(row_index=1, should_refresh=True, should_scroll=True)
    assert rows[1][STATUS_COL] == STATUS_PROCESSING

    complete = apply_grid_update(rows, GridUpdate("complete", 1, (None, "3.500", STATUS_DONE)))

    assert complete == GridUpdateResult(row_index=1, should_refresh=True, should_scroll=False)
    assert rows[1][DATE_COL] == ""
    assert rows[1][RATE_COL] == "3.500"
    assert rows[1][STATUS_COL] == STATUS_DONE

    error = apply_grid_update(rows, parse_legacy_grid_update(("error", 0, STATUS_ERROR_PROCESSING)))

    assert error == GridUpdateResult(row_index=0, should_refresh=True, should_scroll=False)
    assert rows[0][STATUS_COL] == STATUS_ERROR_PROCESSING

    assert apply_grid_update(rows, parse_legacy_grid_update(("processing", 99))) == GridUpdateResult()

    with pytest.raises(ValueError, match="requires update type"):
        parse_legacy_grid_update(("processing",))


def test_excel_io_loads_and_exports_grid_rows(tmp_path):
    source = tmp_path / "source.xlsx"
    pd.DataFrame([{"code": "A001", "name": "Alpha"}]).to_excel(source, index=False)

    rows, missing = load_grid_rows(source)

    assert missing == []
    assert rows == [{CODE_COL: "A001", NAME_COL: "Alpha", DATE_COL: "", RATE_COL: "", STATUS_COL: "대기 중"}]

    rows[0][DATE_COL] = "2026/05/08"
    rows[0][RATE_COL] = "3.5000"
    output = export_grid_rows(rows, tmp_path / "out.xlsx", rate_decimal_places=4)
    with pd.ExcelFile(output) as workbook:
        assert workbook.sheet_names == ["OCR_Results"]

    workbook = load_workbook(output, data_only=True)
    worksheet = workbook["OCR_Results"]
    assert [cell.value for cell in worksheet[1]] == [CODE_COL, NAME_COL, DATE_COL, RATE_COL, STATUS_COL]
    assert worksheet.cell(row=2, column=1).value == "A001"
    assert worksheet.cell(row=2, column=3).value.strftime("%Y/%m/%d") == "2026/05/08"
    assert worksheet.cell(row=2, column=3).number_format == "yyyy/mm/dd"
    assert worksheet.cell(row=2, column=4).value == 3.5
    assert worksheet.cell(row=2, column=4).number_format == "0.0000"


def test_excel_io_uses_configured_rate_decimal_places(tmp_path):
    rows = [
        {CODE_COL: "A001", NAME_COL: "Alpha", DATE_COL: "", RATE_COL: "3.5", STATUS_COL: STATUS_DONE},
    ]

    output = export_grid_rows(rows, tmp_path / "out.xlsx", rate_decimal_places=2)
    worksheet = load_workbook(output, data_only=True)["OCR_Results"]

    assert worksheet.cell(row=2, column=3).value in ("", None)
    assert worksheet.cell(row=2, column=3).number_format == "yyyy/mm/dd"
    assert worksheet.cell(row=2, column=4).value == 3.5
    assert worksheet.cell(row=2, column=4).number_format == "0.00"


def test_excel_io_loads_korean_grid_headers(tmp_path):
    source = tmp_path / "korean.xlsx"
    pd.DataFrame([{CODE_COL: "A001", NAME_COL: "알파"}]).to_excel(source, index=False)

    rows, missing = load_grid_rows(source)

    assert missing == []
    assert rows == [
        {CODE_COL: "A001", NAME_COL: "알파", DATE_COL: "", RATE_COL: "", STATUS_COL: STATUS_WAITING}
    ]


def test_excel_io_loads_blank_cells_as_empty_strings(tmp_path):
    source = tmp_path / "blank_cells.xlsx"
    pd.DataFrame(
        [
            {"code": "A001", "name": None},
            {"code": None, "name": "Blank Code"},
        ]
    ).to_excel(source, index=False)

    rows, missing = load_grid_rows(source)

    assert missing == []
    assert rows[0][NAME_COL] == ""
    assert rows[1][CODE_COL] == ""
    assert rows[0][NAME_COL] != "nan"
    assert rows[1][CODE_COL] != "nan"


def test_table_model_normalizes_nan_values_for_grid_and_export():
    row = {
        CODE_COL: float("nan"),
        NAME_COL: None,
        DATE_COL: "2026/05/08",
        RATE_COL: float("nan"),
        STATUS_COL: STATUS_STOPPED,
    }

    assert grid_cell_text(float("nan")) == ""
    assert row_for_copy(row) == f"\t\t2026/05/08\t\t{STATUS_STOPPED}"
    assert grid_row_values(row) == ("", "", "2026/05/08", "", STATUS_STOPPED)
    assert rows_for_export([row]) == [
        {CODE_COL: "", NAME_COL: "", DATE_COL: "2026/05/08", RATE_COL: "", STATUS_COL: ""}
    ]


def test_excel_io_normalizes_writer_failures(tmp_path):
    rows = [
        {
            CODE_COL: "A001",
            NAME_COL: "Alpha\x01",
            DATE_COL: "2026/05/08",
            RATE_COL: "3.500",
            STATUS_COL: "",
        }
    ]

    with pytest.raises(ExcelIOError, match="could not be written"):
        export_grid_rows(rows, tmp_path / "out.xlsx")
