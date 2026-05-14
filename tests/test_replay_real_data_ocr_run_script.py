from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from PIL import Image

from checkocr2.models import DATE_COL, RATE_COL, STATUS_COL, STATUS_DONE
from scripts.replay_real_data_ocr_run import replay_real_data_ocr_run


class FakeReader:
    def __init__(self, outputs):
        self.outputs = iter(outputs)
        self.calls = []

    def readtext(self, image, detail=0, **kwargs):
        self.calls.append({"detail": detail, "kwargs": kwargs})
        return [next(self.outputs)]


class FailingReader:
    def readtext(self, image, detail=0, **kwargs):
        raise RuntimeError("ocr failed")


def create_replay_workspace(tmp_path: Path) -> tuple[Path, Path]:
    input_excel = tmp_path / "live_smoke_input.xlsx"
    pd.DataFrame([{"code": "KRTEST001", "name": "Alpha"}]).to_excel(input_excel, index=False)

    day_dir = tmp_path / "20260513"
    image_dir = day_dir / "images"
    image_dir.mkdir(parents=True)
    Image.new("RGB", (100, 80), "white").save(image_dir / "KRTEST001.png")
    return input_excel, day_dir


def test_replay_real_data_ocr_run_writes_workbook_and_report(tmp_path):
    input_excel, day_dir = create_replay_workspace(tmp_path)
    output_dir = tmp_path / "replay"
    reader = FakeReader(["2026-05-13", "3.500"])

    summary = replay_real_data_ocr_run(
        input_excel=input_excel,
        day_dir=day_dir,
        output_dir=output_dir,
        layout="cropped",
        date_roi="0,0,0.5,0.5",
        rate_roi="0.5,0,1,0.5",
        upscale_factor=1.0,
        allowlist_mode="field",
        allow_unsafe_output=True,
        reader=reader,
    )

    assert summary["status"] == "ready"
    assert summary["execution_mode"] == "real_data_replay"
    assert summary["processed_count"] == 1
    workbook_path = Path(summary["output_workbook"])
    report_path = Path(summary["run_report"])
    assert workbook_path.exists()
    assert report_path.exists()

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    values = list(workbook.active.iter_rows(values_only=True))
    header = list(values[0])
    row = dict(zip(header, values[1], strict=True))
    assert row[DATE_COL] == "2026/05/13"
    assert row[RATE_COL] == "3.500"
    assert row[STATUS_COL] == STATUS_DONE

    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report["execution_mode"] == "real_data_replay"
    assert report["settings"]["engine"] == "easyocr"
    assert report["summary"]["processed_count"] == 1
    assert report["rows"][0]["date"] == "2026/05/13"
    assert report["rows"][0]["rate"] == "3.500"
    assert reader.calls[0]["kwargs"]["allowlist"] is not None


def test_replay_real_data_ocr_run_uses_pipeline_reader_for_rate_fields(
    tmp_path,
    monkeypatch,
):
    import scripts.replay_real_data_ocr_run as replay_module

    input_excel, day_dir = create_replay_workspace(tmp_path)
    output_dir = tmp_path / "replay"
    crop_reader = FakeReader(["2026-05-13"])

    class PipelineReader:
        def __init__(self):
            self.calls = []

        def readtext(self, image, detail=0, **kwargs):
            self.calls.append({"detail": detail, "kwargs": kwargs})
            return ["3.500"]

    pipeline_reader = PipelineReader()
    pipeline_calls = []

    def fake_create_pipeline_reader(languages, *, gpu=False):
        pipeline_calls.append((list(languages), gpu))
        return pipeline_reader

    monkeypatch.setattr(
        replay_module,
        "create_paddleocr_pipeline_reader",
        fake_create_pipeline_reader,
    )

    summary = replay_real_data_ocr_run(
        input_excel=input_excel,
        day_dir=day_dir,
        output_dir=output_dir,
        engine="paddle",
        layout="cropped",
        date_roi="0,0,0.5,0.5",
        rate_roi="0.5,0,1,0.5",
        upscale_factor=1.0,
        allowlist_mode="field",
        allow_unsafe_output=True,
        reader=crop_reader,
    )

    assert summary["status"] == "ready"
    assert summary["accepted"] is True
    assert pipeline_calls == [(["ko", "en"], False)]
    assert len(crop_reader.calls) == 1
    assert len(pipeline_reader.calls) == 1
    assert crop_reader.calls[0]["kwargs"]["allowlist"] == "0123456789./-"
    assert pipeline_reader.calls[0]["kwargs"]["allowlist"] == "0123456789.,%"

    workbook = load_workbook(Path(summary["output_workbook"]), read_only=True, data_only=True)
    values = list(workbook.active.iter_rows(values_only=True))
    header = list(values[0])
    row = dict(zip(header, values[1], strict=True))
    assert row[DATE_COL] == "2026/05/13"
    assert row[RATE_COL] == "3.500"
    assert row[STATUS_COL] == STATUS_DONE


def test_replay_real_data_ocr_run_fails_closed_on_missing_image(tmp_path):
    input_excel, day_dir = create_replay_workspace(tmp_path)
    (day_dir / "images" / "KRTEST001.png").unlink()

    summary = replay_real_data_ocr_run(
        input_excel=input_excel,
        day_dir=day_dir,
        output_dir=tmp_path / "replay",
        allow_unsafe_output=True,
        reader=FakeReader(["2026-05-13", "3.500"]),
    )

    assert summary["accepted"] is False
    assert summary["status"] == "not_ready"
    assert summary["processed_count"] == 0
    assert summary["missing_image_count"] == 1
    assert "processed_count below row_count: 0 < 1" in summary["errors"]
    assert "missing source images: 1" in summary["errors"]

    report = json.loads(Path(summary["run_report"]).read_text(encoding="utf-8"))
    assert report["errors"]
    assert report["summary"]["processed_count"] == 0


def test_replay_real_data_ocr_run_fails_closed_on_ocr_exception(tmp_path):
    input_excel, day_dir = create_replay_workspace(tmp_path)

    summary = replay_real_data_ocr_run(
        input_excel=input_excel,
        day_dir=day_dir,
        output_dir=tmp_path / "replay",
        allow_unsafe_output=True,
        reader=FailingReader(),
    )

    assert summary["accepted"] is False
    assert summary["status"] == "not_ready"
    assert summary["processing_error_count"] == 1
    assert "row processing errors: 1" in summary["errors"]


def test_replay_real_data_ocr_run_rejects_unsafe_output(tmp_path):
    input_excel, day_dir = create_replay_workspace(tmp_path)

    with pytest.raises(ValueError, match="output_dir must be under"):
        replay_real_data_ocr_run(
            input_excel=input_excel,
            day_dir=day_dir,
            output_dir=Path("docs").resolve(),
            reader=FakeReader(["2026-05-13", "3.500"]),
        )


def test_replay_real_data_ocr_run_cli_reports_missing_inputs(tmp_path):
    result = subprocess.run(
        [
            sys.executable,
            "scripts/replay_real_data_ocr_run.py",
            "--input-excel",
            str(tmp_path / "missing.xlsx"),
            "--day-dir",
            str(tmp_path / "missing-day"),
        ],
        check=False,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 2
    assert "input Excel not found" in result.stderr


def test_replay_real_data_ocr_run_cli_returns_failure_for_incomplete_replay(tmp_path):
    input_excel, day_dir = create_replay_workspace(tmp_path)
    (day_dir / "images" / "KRTEST001.png").unlink()

    result = subprocess.run(
        [
            sys.executable,
            "scripts/replay_real_data_ocr_run.py",
            "--input-excel",
            str(input_excel),
            "--day-dir",
            str(day_dir),
            "--output-dir",
            str(tmp_path / "replay"),
            "--allow-unsafe-output",
        ],
        check=False,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 2
    assert json.loads(result.stdout)["status"] == "not_ready"
